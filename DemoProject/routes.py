from flask import Flask, render_template, request, jsonify, send_file
from textrank import pagerank, sentence_similarity, build_similarity_matrix
from surface_feature import get_surface_score
from nmf import get_grs_score
from rank_sentences import get_top_list
from lexrank import STOPWORDS, LexRank
import re
import nltk
# from nltk.stem import WordNetLemmatizer
from nltk.corpus import stopwords
from nltk.tokenize import sent_tokenize
from timeit import default_timer as timer
import os
import logging
import numpy as np
import openpyxl


# nltk.download('stopwords')
# nltk.download('wordnet')

app = Flask(__name__)


@app.route('/')
def root():
    path=os.path.join('files/','s.txt')
    f = open(path, 'r', encoding='utf-8', errors='ignore')
    text = f.read()
    f.close()

    logging.warning('File Read Successfull')
    start=timer()
    summary = summary_nmf_method(text)
    t=timer()-start
    n=summary[1]
    new_row=list()
    new_row.append(n)
    new_row.append(t)

    path=os.path.join('results/','result.xlsx')
    wb = openpyxl.load_workbook(filename=path)
    ws = wb.get_sheet_by_name('Sheet1')
    row = ws.max_row+ 1

    for i in range(2):
        ws.cell(row=row, column=i+1, value=new_row[i])

    wb.save(path)

    return 'Time Required : '+str(t)+' seconds '+ summary[0]


def summary_nmf_method(text):
    #lemma = WordNetLemmatizer()
    lemma = nltk.wordnet.WordNetLemmatizer()
    # sent_list = re.split('\.|\?|\!', text)
    # sent_list.pop()

    sent_list=sent_tokenize(text)
    i=len(sent_list)-1

    while(i>=0):
        if(len(sent_list[i].split())<3):
            del sent_list[i]
        i-=1

    docs = sent_list.copy()
    n = len(docs)

    logging.warning('No. of sentences')
    logging.warning(str(n))

    for i in range(n):
        if (i == 0):
            docs[i] = docs[i].replace(u'\ufeff', '')
            sent_list[i] = sent_list[i].replace(u'\ufeff', '')
        docs[i] = " ".join(docs[i].split()).lower()
        sent_list[i] = " ".join(sent_list[i].split())

    stop_words = set(stopwords.words('english'))

    for i in range(n):
        sentence = ''
        for w in docs[i].split():
            if w not in stop_words:
                sentence += lemma.lemmatize(w) + ' '
        docs[i] = sentence

    logging.warning('Preprocessing complete')

    GRS_sen = get_grs_score(docs)

    logging.warning('GRS calculated')

    surface_score = get_surface_score(docs)

    logging.warning('Surface Score Calculated')

    # tx = pagerank(docs)
    # txr_score=np.array(tx)
    # maxTex=txr_score.max()
    # txr_score=(100*txr_score)/maxTex

    logging.warning('Pagerank Calculated')

    lxr = LexRank(docs)
    lx = lxr.rank_sentences(docs, threshold=None, fast_power_method=True)

    lxr_score=np.array(lx)
    maxLex=lxr_score.max()
    lxr_score=(100*lxr_score)/maxLex

    logging.warning('Lexrank Calculated')

    total_score = []

    for i in range(n):
        t_sum = float(GRS_sen[i]) + float(surface_score[i]) + float(lxr_score[i])
        total_score.append(t_sum)

    logging.warning('Total score Calculated')

    copy_score = total_score.copy()
    top_list = get_top_list(copy_score)

    logging.warning('TopList Ready !!!')

    # summary_final='<h3>Total Sentecnes :'+str(len(total_score))+'</h3>'+'<h3>Sentecnes Selected :'+str(len(top_list))+'</h3>'
    # i=0
    # while(i<n):
    #     if total_score[i] in top_list:
    #         summary_final+='<p>'+sent_list[i]+'&nbsp; &nbsp;'+str(total_score[i])+'</p>'
    #     else:
    #         summary_final+='<p>----- '+sent_list[i]+'&nbsp; &nbsp;'+str(total_score[i])+' -----</p>'
    #     i+=1

    summary_final = '<h3>Total Sentecnes :' + str(len(total_score)) + '</h3>'

    for i in range(n):
        if (total_score[i] in top_list):
            summary_final += '<p style="color:#00ff00">' + sent_list[i] +'<br>'+str(total_score[i])+ '</p>'
        else:
            summary_final += '<p style="color:#ff0000">' + sent_list[i] +'<br>'+str(total_score[i])+ '</p>'

    result=list()
    result.append(summary_final);
    result.append(n)
    return result


if __name__ == '__main__':
    app.run(port=8000, debug=True)
